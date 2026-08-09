import os
from typing import List
import openpyxl
from src.writers.base import DocumentWriter
from src.fields.models import FormField

class XlsxWriter(DocumentWriter):
    """
    Writer for Excel (XLSX) spreadsheets.
    Mutates cell values in-place, preserving cell borders, fonts, formulas, and formatting.
    Handles merged cells by resolving to the anchor cell before writing.
    """

    def _resolve_merged_cell(self, sheet, row: int, column: int):
        """
        If the target cell is a MergedCell (read-only), find and return the
        anchor cell of the merged range that contains it.
        """
        cell = sheet.cell(row=row, column=column)
        # Lazy import to remain compatible with test-time module mocking
        try:
            from openpyxl.cell.cell import MergedCell as _MergedCell
        except (ImportError, AttributeError):
            # If openpyxl is mocked in tests, skip merged-cell handling
            return cell
        if isinstance(cell, _MergedCell):
            for merged_range in sheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    return sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
            # Should not happen, but fall back to original cell reference
            return cell
        return cell

    def fill(self, source_path: str, fields: List[FormField], output_path: str) -> None:
        if not os.path.exists(source_path):
            raise FileNotFoundError(f"XLSX template not found: {source_path}")
            
        # Load workbook while keeping styles and formulas (data_only=False)
        wb = openpyxl.load_workbook(source_path, data_only=False)
        
        for field in fields:
            if field.answer is None:
                continue
                
            sheet_name = field.metadata.get("sheet_name")
            if sheet_name and sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
            elif field.table_index is not None and field.table_index < len(wb.worksheets):
                sheet = wb.worksheets[field.table_index]
            else:
                continue
                
            # Resolve merged cells to their anchor before writing
            cell = self._resolve_merged_cell(sheet, field.row, field.column)
            cell.value = str(field.answer)
            
        wb.save(output_path)
