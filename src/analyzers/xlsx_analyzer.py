import os
import re
from typing import List, Set
import openpyxl
from src.analyzers.base import AbstractAnalyzer
from src.fields.models import FormField, FieldType

class XlsxAnalyzer(AbstractAnalyzer):
    """
    Layout analyzer for Excel (XLSX) spreadsheets.
    Detects empty input cells aligned horizontally or vertically with labels.
    """

    def clean_label(self, text: str) -> str:
        text = str(text).strip()
        text = re.sub(r'[:：_．\.\-\s]+$', '', text)
        return text.strip()

    def is_label(self, val) -> bool:
        """
        Determines if a cell value represents a form label.
        Labels are non-empty strings that are not numeric and do not start with '='.
        """
        if val is None:
            return False
        val_str = str(val).strip()
        if not val_str:
            return False
        # Skip formulas and numbers
        if val_str.startswith("="):
            return False
        if re.match(r'^\d+(\.\d+)?$', val_str):
            return False
        return True

    def analyze(self, file_path: str) -> List[FormField]:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"XLSX template not found: {file_path}")
            
        # Load workbook, keep formulas (data_only=False) to preserve layout structure
        wb = openpyxl.load_workbook(file_path, data_only=False)
        fields = []
        
        for sheet_idx, sheet in enumerate(wb.worksheets):
            sheet_name = sheet.title
            visited: Set[str] = set()
            
            # Scan max 100 rows and 50 columns to optimize search and prevent timeouts
            max_row = min(sheet.max_row, 100)
            max_col = min(sheet.max_column, 50)
            
            for r in range(1, max_row + 1):
                for c in range(1, max_col + 1):
                    cell = sheet.cell(row=r, column=c)
                    val = cell.value
                    
                    if not self.is_label(val):
                        continue
                        
                    label = self.clean_label(val)
                    if not label:
                        continue
                        
                    # Check Case A: Horizontal alignment (input to the right of label)
                    if c < max_col:
                        right_cell = sheet.cell(row=r, column=c + 1)
                        coord = right_cell.coordinate
                        if right_cell.value is None and coord not in visited:
                            fields.append(FormField(
                                id=f"xlsx_{sheet_name}_{coord}",
                                label=label,
                                field_type=FieldType.TEXT,
                                table_index=sheet_idx,
                                row=r,
                                column=c + 1,
                                metadata={
                                    "sheet_name": sheet_name,
                                    "cell_coordinate": coord
                                }
                            ))
                            visited.add(coord)
                            continue
                            
                    # Check Case B: Vertical alignment (input below label)
                    if r < max_row:
                        below_cell = sheet.cell(row=r + 1, column=c)
                        coord = below_cell.coordinate
                        if below_cell.value is None and coord not in visited:
                            fields.append(FormField(
                                id=f"xlsx_{sheet_name}_{coord}",
                                label=label,
                                field_type=FieldType.TEXT,
                                table_index=sheet_idx,
                                row=r + 1,
                                column=c,
                                metadata={
                                    "sheet_name": sheet_name,
                                    "cell_coordinate": coord
                                }
                            ))
                            visited.add(coord)
                            
        return fields
